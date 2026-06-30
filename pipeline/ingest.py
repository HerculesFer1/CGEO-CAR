#!/usr/bin/env python3
"""
CGEO-CAR Dashboard - Ingest Helper
==================================
Le o Excel mensal do SICAR-PI e atualiza data/data.json automaticamente.

Le do Excel (sheet "Planilha1" ou raw "Relatorio-Buscar-Imoveis"):
  - Contagens por Fase do Processo (AGUARDANDO GESTOR, CANCELADOS, PENDENTE, VALIDADOS)
  - Contagens por Situacao do Imovel (Ativo, Cancelado, Pendente, Suspenso, Retificado)

Le do CSV ranking_uf.csv (UF;Total):
  - Total nacional por UF

Uso:
  python ingest.py --excel "C:/path/to/Planilha_Secretario.xlsx" \\
                   --uf-csv "data/uf_ranking_julho.csv" \\
                   --month "Julho" --year 2026 \\
                   --prev-month-short "Jun/26"

Setor: CGEO / SEMARH-PI
"""
import argparse
import csv
import json
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")

MONTH_ABBR = {
    "Janeiro":"Jan","Fevereiro":"Fev","Marco":"Mar","Marco":"Mar","Abril":"Abr",
    "Maio":"Mai","Junho":"Jun","Julho":"Jul","Agosto":"Ago",
    "Setembro":"Set","Outubro":"Out","Novembro":"Nov","Dezembro":"Dez"
}

def read_pi_from_excel(xlsx_path):
    """Le a Planilha1 (pivot) ou agrega da sheet raw."""
    try:
        import openpyxl
    except ImportError:
        print("ERRO: 'openpyxl' nao instalado. Rode: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Tenta primeiro a Planilha1 (pivot pronto)
    if "Planilha1" in wb.sheetnames:
        ws = wb["Planilha1"]
        rows = list(ws.iter_rows(values_only=True))
        fases, situacoes = {}, {}
        mode = None
        for r in rows:
            if r[0] and isinstance(r[0], str):
                t = r[0].strip()
                if t.startswith(("Rotulos","R" )) and "Fase" in (r[1] or ""):
                    mode = "fase"; continue
                if t.startswith(("Rotulos","R")) and "Situa" in (r[1] or ""):
                    mode = "sit"; continue
                if mode and r[1] is not None and t != "Total Geral":
                    if mode == "fase": fases[t] = int(r[1])
                    elif mode == "sit": situacoes[t] = int(r[1])

        if fases and situacoes:
            print(f"  Lido de Planilha1: {len(fases)} fases, {len(situacoes)} situacoes")
            return _build_pi(fases, situacoes)

    # Fallback: agrega da sheet raw
    print("  Planilha1 nao detectada. Agregando da sheet raw...")
    raw_sheet = next((s for s in wb.sheetnames if "Relatorio" in s or "Buscar" in s), None)
    if not raw_sheet:
        raw_sheet = wb.sheetnames[-1]
    ws = wb[raw_sheet]
    fases, situacoes = {}, {}
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0: continue
        if not r or not r[-1]: continue
        fase = r[-1]; sit = r[-2]
        fases[fase] = fases.get(fase, 0) + 1
        situacoes[sit] = situacoes.get(sit, 0) + 1
    print(f"  Agregado: {sum(fases.values())} registros, {len(fases)} fases")
    return _build_pi(fases, situacoes)

def _build_pi(fases, situacoes):
    return {
        "total": sum(fases.values()),
        "ag_gestor": fases.get("AGUARDANDO GESTOR", 0),
        "validados": fases.get("VALIDADOS", 0),
        "cancelados": fases.get("CANCELADOS", 0),
        "pendentes": fases.get("PENDENTE", 0),
        "suspensos": situacoes.get("Suspenso", 0),
    }

def read_uf_csv(csv_path):
    """CSV no formato 'UF;Total' ou 'UF,Total'. Cabecalho opcional."""
    ranking = {}
    with open(csv_path, encoding="utf-8") as f:
        # detecta separador
        first = f.readline()
        sep = ";" if ";" in first else ","
        f.seek(0)
        reader = csv.reader(f, delimiter=sep)
        for row in reader:
            if not row or len(row) < 2: continue
            uf = row[0].strip()
            if len(uf) != 2 or not uf.isalpha(): continue  # ignora cabecalho/total
            try:
                val = int(str(row[1]).replace(".","").replace(",","").strip())
            except ValueError:
                continue
            ranking[uf.upper()] = val
    return ranking

def update_data_json(data_path, pi, uf_ranking, month_long, year, prev_short):
    data = json.loads(Path(data_path).read_text(encoding="utf-8"))
    short = f"{MONTH_ABBR[month_long]}/{str(year)[-2:]}"

    # Atualiza header
    data["month_long"] = month_long
    data["month_short"] = short
    data["previous_month_short"] = prev_short
    data["year"] = year
    data["pi"] = {**data["pi"], **pi}
    # Mantem o _doc original
    if "_doc" in data.get("pi", {}):
        data["pi"]["_doc"] = json.loads(Path(data_path).read_text(encoding="utf-8"))["pi"].get("_doc","")
    data["uf_ranking"] = {"_doc": data["uf_ranking"].get("_doc",""), **uf_ranking}

    # Append historico (somente se ainda nao registrado)
    hist = data["history"]
    if hist["labels"][-1] != short:
        hist["labels"].append(short)
        hist["ag_gestor"].append(pi["ag_gestor"])
        hist["validados"].append(pi["validados"])
        hist["cancelados"].append(pi["cancelados"])
        hist["pendentes"].append(pi["pendentes"])
        hist["suspensos"].append(pi["suspensos"])

    # PI ranking dos ultimos 4 meses
    pi_val = uf_ranking.get("PI", 0)
    p4 = data["pi_analyses_4mo"]
    if p4["labels"][-1] != short:
        p4["labels"] = p4["labels"][1:] + [short]
        p4["pi"] = p4["pi"][1:] + [pi_val]
        total_br = sum(v for v in uf_ranking.values())
        p4["brasil"] = p4["brasil"][1:] + [total_br]

    Path(data_path).write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  data.json atualizado: {data_path}")

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True, help="Caminho do .xlsx do SICAR-PI")
    ap.add_argument("--uf-csv", required=True, help="CSV com UF;Total nacional")
    ap.add_argument("--month", required=True, help="Mes por extenso (ex: Julho)")
    ap.add_argument("--year", type=int, required=True)
    ap.add_argument("--prev-month-short", required=True, help="Mes anterior abreviado (ex: Jun/26)")
    ap.add_argument("--data", default="data/data.json")
    args = ap.parse_args()

    here = Path(__file__).parent
    print(f"-> Lendo Excel: {args.excel}")
    pi = read_pi_from_excel(args.excel)
    print(f"   PI Total: {pi['total']:,}".replace(",","."))
    print(f"   AG: {pi['ag_gestor']:,} | Pend: {pi['pendentes']:,} | Val: {pi['validados']:,} | Canc: {pi['cancelados']:,}".replace(",","."))

    print(f"-> Lendo UF CSV: {args.uf_csv}")
    uf = read_uf_csv(args.uf_csv)
    print(f"   UFs: {len(uf)} | Total Brasil: {sum(uf.values()):,}".replace(",","."))

    update_data_json(here / args.data, pi, uf, args.month, args.year, args.prev_month_short)
    print("OK. Agora rode: python build.py")

if __name__ == "__main__":
    main()
